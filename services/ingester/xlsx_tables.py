"""
xlsx_tables — structural XLSX table extraction for the ragfarm ingester.

Split out of ingester.py so the messy-spreadsheet logic lives in one place and is
unit-testable in isolation. ingester.py imports `iter_xlsx` from here.

Two header MODELS, auto-routed per table region by `_classify_header_model`:

  SINGLE-ROW  (EPC25 family: New DC / Old DC / Stacked / Plain)
    - one physical header row, found by scored detection (_detect_header_row)
    - OR no header at all (Plain) -> bare-value join
    - stacked tables split on a repeated header-shaped row (_is_new_table_header)
    - grouping: dense repetition | sparse carry-forward | vertical merge

  MULTI-ROW   (SA_Hosting family: title band + super-header bands + sub-labels)
    - section title band (A1:X1 'Central Services') spanning the width
    - 2- or 3-row header stack: per-column label is the bottom-most non-blank row,
      prefixed by any horizontal band merge above it (disambiguates the repeated
      IP/FQDN/VLAN/Network triples under app-mgmt / backup / storage bands)
    - a following title band with NO header of its own REUSES the prior header
      ('EPC tenant' inherits 'Central Services' columns)
    - unheaded leading group column (col A merged labels) absorbed into the span

Cross-cutting (both models):
    - trailing-trim: a table ends at the last row whose IDENTITY columns
      (hostname/IP/FQDN/servername) are populated; totals rows ('number of VMs')
      and free-text notes below the data are dropped.
    - identifiers kept verbatim (no lowercasing); int-as-float cleaned
      (603423146.0 -> 603423146) so sparse exact-match works.

Routing decision and every boundary are logged at INFO so a corpus run is auditable.

NOTE on language: the identity-column regex `_ANCHOR_RE` includes Czech terms
(`název hostitele`, `jméno serveru`, `IP adresa`). If a future sheet uses other
Czech identity headers, extend that pattern — otherwise anchor detection falls
back to the leftmost column and trailing notes can leak back in.
"""
import re
import uuid
import logging
from typing import Iterator, Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

log = logging.getLogger("ingester.xlsx")

# --- tunables ---------------------------------------------------------------
HEADER_SCAN_ROWS = 25
MAX_TABLES_PER_SHEET = 50
MAX_HEADER_STACK = 4          # multi-row header: max physical rows in the stack
DATA_PROBE_ROWS = 8           # how far below a header to look for the first data row
TITLE_BAND_FRAC = 0.6         # a horizontal merge >= this * width is a section title
GROUP_COL_FILL_FRAC = 0.6     # a group-label col is populated in < this frac of rows

# identity columns: a real infra row populates at least one of these. Used for the
# trailing-trim terminator. NOT the VM-name column (collides with free-text notes).
_ANCHOR_RE = re.compile(
    r"\b(host\s*name|hostname|server\s*name|servername|ip\s*address|ip\b|fqdn"
    r"|n[áa]zev\s*hostitele|jm[ée]no\s*serveru|ip\s*adresa)\b", re.I)

# tokens that mark a cell as real infra data (used to find the first data row,
# which in turn fixes the bottom of a multi-row header stack)
_FQDN = re.compile(r"\b[a-z0-9-]+(\.[a-z0-9-]+){2,}\b", re.I)
_IP   = re.compile(r"\b\d{1,3}(\.\d{1,3}){3}\b")
_NUM  = re.compile(r"-?\d+(\.\d+)?$")


# ============================================================================
# small cell helpers
# ============================================================================
def clean_cell(v) -> str:
    """Verbatim stringify, fixing openpyxl int-as-float and stray nbsp."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).replace("\xa0", " ").strip()


def _clean_header_label(v) -> str:
    """Header cell text, trimmed and stripped of stray trailing colons only
    (e.g. 'E-mail:' -> 'E-mail'). Other punctuation ('[', ']', '(', ')', etc.)
    is preserved verbatim; extend here if more needs trimming."""
    return str(v).strip().rstrip(":").strip()


def _is_numeric(v) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    return bool(_NUM.match(str(v).strip().replace(",", "")))


def _fill_rgb(cell) -> Optional[str]:
    f = cell.fill
    if f and f.patternType:
        rgb = f.fgColor.rgb
        return rgb if isinstance(rgb, str) else None
    return None


def _row_nonempty_cols(wsv, ri, cmax) -> list:
    return [ci for ci in range(1, cmax + 1)
            if (v := wsv.cell(row=ri, column=ci).value) is not None and str(v).strip() != ""]


def _is_all_blank(wsv, ri, c1, c2) -> bool:
    return all((v := wsv.cell(row=ri, column=ci).value) is None or str(v).strip() == ""
               for ci in range(c1, c2 + 1))


def point_id(*parts) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_URL, ":".join(str(p) for p in parts)))


def _serialize_kv(keys: dict, values: dict, sheet: Optional[str]) -> Optional[str]:
    parts = []
    if sheet:
        parts.append(f"sheet: {sheet}")
    for ci, k in keys.items():
        val = clean_cell(values.get(ci))
        if val == "":
            continue
        parts.append(f"{k}: {val}")
    return ", ".join(parts) if len(parts) > (1 if sheet else 0) else None


def _serialize_bare(wsv, ri, c1, c2, sheet) -> Optional[str]:
    """Headerless row -> bare value join, identifiers verbatim, no synthetic keys."""
    vals = [clean_cell(wsv.cell(row=ri, column=ci).value) for ci in range(c1, c2 + 1)]
    vals = [v for v in vals if v != ""]
    if not vals:
        return None
    head = f"sheet: {sheet} | " if sheet else ""
    return head + " | ".join(vals)


# ============================================================================
# identity / data-row tests (shared terminator logic)
# ============================================================================
def _anchor_cols(keys: dict) -> list:
    """Identity columns by header name; fall back to leftmost key column."""
    a = [ci for ci, k in keys.items() if k != "group" and _ANCHOR_RE.search(str(k))]
    if not a and keys:
        a = [min(keys)]
    return a


def _row_is_data(wsv, ri, anchors) -> bool:
    return any((v := wsv.cell(row=ri, column=ci).value) is not None and str(v).strip() != ""
               for ci in anchors)


def _last_data_row(wsv, anchors, hdr_bottom, end) -> int:
    """Highest row in (hdr_bottom, end] whose identity cols are populated. Rows
    below it are trailing totals/notes -> excluded. Interior gaps are preserved
    because we trim from the BOTTOM only."""
    last = hdr_bottom
    for rr in range(hdr_bottom + 1, end + 1):
        if _row_is_data(wsv, rr, anchors):
            last = rr
    return last


def _row_has_identity_token(wsv, ri, cmax) -> bool:
    for c in range(1, cmax + 1):
        v = wsv.cell(row=ri, column=c).value
        if v is None:
            continue
        s = str(v)
        if _FQDN.search(s) or _IP.search(s):
            return True
    return False


def _first_data_row(wsv, top, cmax, limit=DATA_PROBE_ROWS) -> Optional[int]:
    for r in range(top, min(top + limit, wsv.max_row) + 1):
        if _row_has_identity_token(wsv, r, cmax):
            return r
    return None


# ============================================================================
# merges: horizontal bands (super-headers) and vertical ranges (grouping)
# ============================================================================
def _hbands(wss, row) -> dict:
    """col -> label for horizontal merges on `row` (super-header bands)."""
    out = {}
    for mr in wss.merged_cells.ranges:
        a, r1, b, r2 = range_boundaries(str(mr))
        if r1 == r2 == row and b > a:
            v = wss.cell(row=r1, column=a).value
            if v and str(v).strip():
                for ci in range(a, b + 1):
                    out[ci] = str(v).strip()
    return out


def _title_band(wss, ri, cmax) -> Optional[str]:
    """A single horizontal merge spanning most of the width = section title row."""
    for mr in wss.merged_cells.ranges:
        a, r1, b, r2 = range_boundaries(str(mr))
        if r1 == r2 == ri and (b - a + 1) >= TITLE_BAND_FRAC * cmax:
            v = wss.cell(row=r1, column=a).value
            if v and str(v).strip():
                return str(v).strip()
    return None


def _vmerge_map(wss, c1, c2, row_lo, row_hi) -> dict:
    """col -> {row: value} for VERTICAL merged ranges, clamped to [row_lo,row_hi]
    so a merge in one region never leaks into the next."""
    m = {}
    for mr in wss.merged_cells.ranges:
        a, r1, b, r2 = range_boundaries(str(mr))
        if a == b and r2 > r1 and c1 <= a <= c2:
            lo, hi = max(r1, row_lo), min(r2, row_hi)
            if lo > hi:
                continue
            val = wss.cell(row=r1, column=a).value
            if val is None or str(val).strip() == "":
                continue
            m.setdefault(a, {})
            for ri in range(lo, hi + 1):
                m[a][ri] = val
    return m


# ============================================================================
# SINGLE-ROW header model (EPC25 family)
# ============================================================================
def _detect_header_row(wsv, wss, start_ri, cmax, scan=HEADER_SCAN_ROWS) -> int:
    """Scored header detection. A header is recognized by being more DISTINCT than
    its neighbours (fill/bold edge), textual, short, and spanning the data width.
    Intra-row formatting uniformity is deliberately NOT required (real headers
    carry random bold/font noise)."""
    rmax = wsv.max_row or start_ri
    hi = min(start_ri + scan, rmax)
    dens = {ri: len(_row_nonempty_cols(wsv, ri, cmax))
            for ri in range(start_ri, min(hi + 5, rmax) + 1)}
    maxdens = max(dens.values()) if dens else 0
    best_ri, best = start_ri, -1.0
    for ri in range(start_ri, hi + 1):
        cols = _row_nonempty_cols(wsv, ri, cmax)
        if not cols:
            continue
        n = len(cols)
        vals = [wsv.cell(row=ri, column=ci).value for ci in cols]
        nbold = sum(1 for ci in cols if wss.cell(row=ri, column=ci).font.bold)
        nstr = sum(1 for v in vals if not _is_numeric(v))
        nshort = sum(1 for v in vals if len(str(v).strip()) <= 40)

        def _distinct(other_ri):
            d = 0
            for ci in cols:
                a = wss.cell(row=ri, column=ci)
                b = wss.cell(row=other_ri, column=ci)
                if _fill_rgb(a) != _fill_rgb(b) or bool(a.font.bold) != bool(b.font.bold):
                    d += 1
            return d / n

        dist_below = _distinct(ri + 1) if ri + 1 <= rmax else 1.0
        dist_above = _distinct(ri - 1) if ri - 1 >= 1 else 1.0
        dmatch = maxdens > 0 and n >= 0.6 * maxdens
        score = (2 * (nstr / n) + (nshort / n)
                 + 1.5 * dist_below + 0.8 * dist_above
                 + 0.5 * (nbold / n) + (2 if dmatch else 0))
        if score > best:
            best_ri, best = ri, score
    return best_ri


def _header_span(wsv, header_ri, cmax) -> tuple:
    cols = _row_nonempty_cols(wsv, header_ri, cmax)
    return (min(cols), max(cols)) if cols else (1, 1)


def _is_new_table_header(wsv, wss, ri, c1, c2) -> bool:
    """Boundary trigger for directly-stacked tables: a header-SHAPED row —
    formatting (bold/fill) AND text-or-blank only (no numbers) AND text>blank."""
    cells = [wsv.cell(row=ri, column=ci).value for ci in range(c1, c2 + 1)]
    ntext = sum(1 for v in cells if v is not None and str(v).strip() != "" and not _is_numeric(v))
    nnum = sum(1 for v in cells if v is not None and str(v).strip() != "" and _is_numeric(v))
    nblank = sum(1 for v in cells if v is None or str(v).strip() == "")
    if nnum > 0:
        return False
    if ntext <= nblank:
        return False
    nbold = sum(1 for ci in range(c1, c2 + 1) if wss.cell(row=ri, column=ci).font.bold)
    nfill = sum(1 for ci in range(c1, c2 + 1) if _fill_rgb(wss.cell(row=ri, column=ci)) is not None)
    return (nbold >= ntext * 0.5) or (nfill >= ntext * 0.5)


def _is_band_boundary(wsv, wss, ri, c1, c2) -> bool:
    """Boundary trigger for a super-header BAND row that reopens a new labeled
    sub-table (e.g. 'Projektový tým' after 'Řízení projektu' in a contact
    list): a blank separator row directly above it, and a horizontal merge
    band starting within [c1, c2] on this row. Catches band rows too sparse
    (mostly blank around the one merged label) to pass `_is_new_table_header`'s
    text>blank test."""
    if ri <= 1 or not _is_all_blank(wsv, ri - 1, c1, c2):
        return False
    bands = _hbands(wss, ri)
    return any(c1 <= ci <= c2 for ci in bands)


def _merged_bands_single(wss, header_ri) -> dict:
    """Single horizontal merge one row above the header (duplicate-column
    disambiguation for the single-row model)."""
    return _hbands(wss, header_ri - 1)


def _build_header_single(wsv, wss, header_ri, c1, c2) -> dict:
    bands = _merged_bands_single(wss, header_ri)
    keys, seen = {}, {}
    for ci in range(c1, c2 + 1):
        v = wsv.cell(row=header_ri, column=ci).value
        if v is None or str(v).strip() == "":
            continue
        k = _clean_header_label(v)
        if ci in bands and bands[ci] != k:
            k = f"{bands[ci]} {k}"
        if k in seen:
            seen[k] += 1; k = f"{k} #{seen[k]}"
        else:
            seen[k] = 1
        keys[ci] = k
    return keys


def _cell_style(c) -> tuple:
    try:
        color = c.font.color.rgb if (c.font.color and isinstance(c.font.color.rgb, str)) else None
    except Exception:
        color = None
    fill = None
    if c.fill and c.fill.patternType and isinstance(c.fill.fgColor.rgb, str):
        fill = c.fill.fgColor.rgb
    return (bool(c.font.bold), bool(c.font.italic), color, fill)


def _ff_candidate_cols(wsv, wss, keys, header_ri, end) -> set:
    """Sparse carry-forward (encoding 2): restricted to the FIRST column, and only
    when it is a VISUALLY DISTINGUISHED group-label column (formatting differs from
    col2). Prevents smearing a one-off value (e.g. 'Application Name') down rows."""
    cols = sorted(keys)
    if len(cols) < 2:
        return set()
    col1, col2 = cols[0], cols[1]
    probe = None
    for rr in range(header_ri + 1, end + 1):
        v = wsv.cell(row=rr, column=col1).value
        if v is not None and str(v).strip() != "":
            probe = rr
            break
    if probe is None:
        return set()
    if _cell_style(wss.cell(row=probe, column=col1)) != _cell_style(wss.cell(row=probe, column=col2)):
        return {col1}
    return set()


def _reuse_trailing_cols(wsv, keys, c2, header_ri, end, band_label, prev_single) -> tuple:
    """Extend a single-row header with TRAILING columns (right of this header's
    own span) that a PRIOR single-header table earlier in the same sheet
    already labeled, and that still carry data in this sub-table's body, but
    have no label of their own in this header row. Covers a repeated
    super-header band whose trailing columns aren't re-typed under the new
    band (e.g. 'Tel:'/'E-mail:' shown once under 'Řízení projektu', not
    repeated under a later 'Projektový tým'). Reused labels are prefixed with
    THIS header's own band, matching directly-labeled columns. Returns the
    (possibly extended) keys dict and c2."""
    if not band_label or not prev_single or not prev_single.get("keys"):
        return keys, c2
    extra = {}
    for ci, label in prev_single["keys"].items():
        if ci <= c2 or ci in keys:
            continue
        has_data = any(
            (v := wsv.cell(row=rr, column=ci).value) is not None and str(v).strip() != ""
            for rr in range(header_ri + 1, end + 1))
        if has_data:
            extra[ci] = label if label.startswith(band_label) else f"{band_label} {label}"
    if not extra:
        return keys, c2
    return {**keys, **extra}, max(c2, max(extra))


def _looks_headerless(wsv, wss, header_ri, c1, c2, end) -> bool:
    """The scored 'header' is bogus if it is shape-identical to the row below AND
    has no formatting edge (Plain: every row identical)."""
    below = header_ri + 1
    if below > end:
        return True

    def shape(ri):
        out = []
        for ci in range(c1, c2 + 1):
            v = wsv.cell(row=ri, column=ci).value
            out.append("_" if v is None or str(v).strip() == ""
                       else "n" if _is_numeric(v) else "t")
        return out

    if shape(header_ri) != shape(below):
        return False
    edge = tot = 0
    for ci in range(c1, c2 + 1):
        a = wss.cell(row=header_ri, column=ci); b = wss.cell(row=below, column=ci)
        tot += 1
        if _fill_rgb(a) != _fill_rgb(b) or bool(a.font.bold) != bool(b.font.bold):
            edge += 1
    return (edge / max(tot, 1)) < 0.3


# ============================================================================
# MULTI-ROW header model (SA_Hosting family)
# ============================================================================
def _compose_header_multi(wsv, wss, top, bottom, cmax) -> dict:
    """Per column, bottom-most non-blank label across the header stack, prefixed by
    any band merge above it. Collapses a 2-/3-row boxed header into flat keys and
    disambiguates repeated sub-labels (IP/FQDN/...) by their super-header band."""
    keys, seen = {}, {}
    bands = {r: _hbands(wss, r) for r in range(top, bottom + 1)}
    for c in range(1, cmax + 1):
        label = None
        for r in range(bottom, top - 1, -1):
            v = wsv.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                label = _clean_header_label(v)
                break
        if label is None:
            continue
        prefix = [bands[r][c] for r in range(top, bottom + 1)
                  if c in bands[r] and bands[r][c] != label]
        k = " ".join(prefix + [label]) if prefix else label
        if k in seen:
            seen[k] += 1; k = f"{k} #{seen[k]}"
        else:
            seen[k] = 1
        keys[c] = k
    return keys


def _extend_for_group_col(wsv, wss, c1, header_bottom, end) -> tuple:
    """Absorb an unheaded leading group-label column (blank header cell, but vertical
    merges or sparse labels below). Returns (new_c1, synth_col or None)."""
    cc1, synth, probe = c1, None, c1 - 1
    while probe >= 1:
        hv = wsv.cell(row=header_bottom, column=probe).value
        if hv is not None and str(hv).strip() != "":
            break
        has_vmerge = any(
            a == b == probe and r2 > r1 and r1 <= end and r2 > header_bottom
            for a, r1, b, r2 in (range_boundaries(str(m)) for m in wss.merged_cells.ranges))
        if not has_vmerge:
            nb = sum(1 for rr in range(header_bottom + 1, end + 1)
                     if (v := wsv.cell(row=rr, column=probe).value) is not None and str(v).strip() != "")
            rows = end - header_bottom
            if not (0 < nb < max(2, rows * GROUP_COL_FILL_FRAC)):
                break
        cc1, synth, probe = probe, probe, probe - 1
    return cc1, synth


# ============================================================================
# per-region classifier + emitters
# ============================================================================
def _classify_header_model(wss, region_top, cmax) -> str:
    """'multi' if the region opens with a title band or a horizontal super-header
    band stack; else 'single'. Cheap structural test, no data inspection."""
    if _title_band(wss, region_top, cmax):
        return "multi"
    # a non-title horizontal band in the first two rows => stacked header
    for r in (region_top, region_top + 1):
        for ci, _ in _hbands(wss, r).items():
            # band that is NOT a full-width title => super-header
            return "multi"
    return "single"


def _emit_single(wsv, wss, sheet, cmax, source_file, region_top, region_end, tnum, report=None,
                  prev_single=None) -> Iterator[dict]:
    """One single-row-header table starting at/after region_top. Yields records and,
    as the LAST yielded item, a control dict {'_next_ri': N} for the caller.

    If `report` is a list, and a heading is actually captured (non-Plain), a meta
    dict is appended to it: sheet/table/model/header position/col span/keys, with
    'rows' filled in once the data rows have been counted.

    `prev_single`, if a dict, carries {'keys'} from the previous single-header
    table in the same sheet, for `_reuse_trailing_cols` and is updated with
    this table's keys once captured (see `_reuse_trailing_cols`)."""
    header_ri = _detect_header_row(wsv, wss, region_top, cmax)
    c1, c2 = _header_span(wsv, header_ri, cmax)

    # provisional end = next new-header-shaped or new-band row (stacked table) within region
    end = region_end
    scan = header_ri + 1
    while scan <= region_end:
        if _is_new_table_header(wsv, wss, scan, c1, c2) or _is_band_boundary(wsv, wss, scan, c1, c2):
            end = scan - 1
            break
        scan += 1

    # headerless region (Plain) -> bare-value join
    if _looks_headerless(wsv, wss, header_ri, c1, c2, end):
        cc1, cc2, any_row = cmax, 1, False
        for rr in range(region_top, end + 1):
            cols = _row_nonempty_cols(wsv, rr, cmax)
            if cols:
                cc1, cc2, any_row = min(cc1, min(cols)), max(cc2, max(cols)), True
        if any_row:
            log.info("%s[%s] t#%d SINGLE/headerless span=%d..%d rows=%d",
                     source_file, sheet, tnum, cc1, cc2, end - region_top + 1)
            ridx = 0
            for rr in range(region_top, end + 1):
                if _is_all_blank(wsv, rr, cc1, cc2):
                    continue
                ridx += 1
                text = _serialize_bare(wsv, rr, cc1, cc2, sheet)
                if text:
                    yield _rec(source_file, sheet, tnum, ridx, text, header_row=None)
        yield {"_next_ri": end + 1}
        return

    keys = _build_header_single(wsv, wss, header_ri, c1, c2)
    if not keys:
        yield {"_next_ri": region_end + 1}
        return
    band_label = next(iter(_merged_bands_single(wss, header_ri).values()), None)
    keys, c2 = _reuse_trailing_cols(wsv, keys, c2, header_ri, end, band_label, prev_single)
    anchors = _anchor_cols(keys)
    data_end = _last_data_row(wsv, anchors, header_ri, end)
    ff_cols = _ff_candidate_cols(wsv, wss, keys, header_ri, data_end)
    vm = _vmerge_map(wss, c1, c2, header_ri + 1, data_end)
    log.info("%s[%s] t#%d SINGLE hdr R%d span=%d..%d cols=%d anchors=%s ff=%d vm=%s rows=%d trim=%d",
             source_file, sheet, tnum, header_ri, c1, c2, len(keys), anchors,
             len(ff_cols), sorted(vm.keys()), data_end - header_ri, end - data_end)

    meta = None
    if report is not None:
        meta = {"sheet": sheet, "table": tnum, "model": "single",
                "header_row": header_ri, "col_span": (c1, c2),
                "keys": list(keys.values()), "rows": 0}
        report.append(meta)

    carry, ridx = {}, 0
    for rr in range(header_ri + 1, data_end + 1):
        if _is_all_blank(wsv, rr, c1, c2):
            carry = {}
            continue
        values = {ci: wsv.cell(row=rr, column=ci).value for ci in keys}
        for ci in keys:
            if ci in vm and rr in vm[ci]:
                cur = values.get(ci)
                if cur is None or str(cur).strip() == "":
                    values[ci] = vm[ci][rr]
        for ci in ff_cols:
            v = values.get(ci)
            if v is not None and str(v).strip() != "":
                carry[ci] = v
            elif ci in carry:
                values[ci] = carry[ci]
        if not _row_is_data(wsv, rr, anchors):
            # interior non-data row: keep only if it set a carry (grouping label)
            if not any(ci in ff_cols for ci in _row_nonempty_cols(wsv, rr, cmax)):
                continue
        ridx += 1
        text = _serialize_kv(keys, values, sheet)
        if text:
            yield _rec(source_file, sheet, tnum, ridx, text, header_row=header_ri)
    if meta is not None:
        meta["rows"] = ridx
    if prev_single is not None:
        prev_single["keys"] = keys
    yield {"_next_ri": end + 1}


def _emit_multi(wsv, wss, sheet, cmax, source_file, region_top, tnum, prev, report=None) -> Iterator[dict]:
    """One multi-row-header table (or a title-band section reusing prev header).
    `prev` carries {'keys','c1','c2'} from the previous multi table for reuse.

    If `report` is a list, a meta dict (position, keys, row count) is appended —
    see `_emit_single` for the shape."""
    rmax = wsv.max_row or 1
    ri = region_top
    title = None
    while ri <= rmax and _title_band(wss, ri, cmax):
        title = _title_band(wss, ri, cmax)
        ri += 1
    if ri > rmax:
        yield {"_next_ri": rmax + 1}
        return

    fdr = _first_data_row(wsv, ri, cmax)
    reused = False
    if fdr is None:
        yield {"_next_ri": ri + 1}
        return
    if fdr == ri and prev.get("keys"):
        # title band immediately followed by data -> reuse previous header
        top = bottom = ri - 1
        keys = dict(prev["keys"]); c1, c2 = prev["c1"], prev["c2"]; reused = True
    else:
        top, bottom = ri, fdr - 1
        keys = _compose_header_multi(wsv, wss, top, bottom, cmax)
        if not keys:
            yield {"_next_ri": fdr + 1}
            return
        c1, c2 = min(keys), max(keys)

    # region end = next title band, else end of sheet
    end = rmax
    scan = fdr
    while scan <= rmax:
        if _title_band(wss, scan, cmax):
            end = scan - 1
            break
        scan += 1

    # absorb unheaded leading group column (both fresh and reused header paths)
    new_c1, synth = _extend_for_group_col(wsv, wss, c1, bottom, end)
    if synth is not None and synth not in keys:
        keys = {synth: "group", **keys}
        c1 = new_c1

    anchors = _anchor_cols(keys)
    data_end = _last_data_row(wsv, anchors, bottom, end)
    vm = _vmerge_map(wss, c1, c2, bottom + 1, data_end)
    log.info("%s[%s] t#%d MULTI title=%r hdr R%d-%d span=%d..%d cols=%d reuse=%s synth=%s anchors=%s vm=%s rows=%d trim=%d",
             source_file, sheet, tnum, title, top, bottom, c1, c2, len(keys),
             reused, synth, anchors, sorted(vm.keys()), data_end - bottom, end - data_end)

    meta = None
    if report is not None:
        meta = {"sheet": sheet, "table": tnum, "model": "multi",
                "header_row": (top, bottom), "col_span": (c1, c2),
                "keys": list(keys.values()), "reused": reused, "rows": 0}
        report.append(meta)

    ridx = 0
    for rr in range(bottom + 1, data_end + 1):
        if _is_all_blank(wsv, rr, c1, c2):
            continue
        values = {ci: wsv.cell(row=rr, column=ci).value for ci in keys}
        for ci in keys:
            if ci in vm and rr in vm[ci]:
                cur = values.get(ci)
                if cur is None or str(cur).strip() == "":
                    values[ci] = vm[ci][rr]
        if not _row_is_data(wsv, rr, anchors):
            continue
        ridx += 1
        text = _serialize_kv(keys, values, sheet)
        if text:
            yield _rec(source_file, sheet, tnum, ridx, text, header_row=bottom)

    if meta is not None:
        meta["rows"] = ridx

    # expose header for reuse by a following title-band section
    if not reused:
        prev["keys"] = {k: v for k, v in keys.items() if v != "group"}
        prev["c1"], prev["c2"] = c1, c2
    yield {"_next_ri": end + 1}


def _rec(source_file, sheet, tnum, ridx, text, header_row) -> dict:
    return {
        "id": point_id(source_file, sheet, tnum, ridx),
        "text": text,
        "payload": {
            "source_file": source_file, "sheet": sheet, "table": tnum,
            "row_index": ridx, "header_row": header_row,
            "kind": "table_row", "lang": "n/a", "text": text,
        },
    }


# ============================================================================
# sheet + workbook drivers
# ============================================================================
def _iter_sheet_tables(wsv, wss, sheet, cmax, source_file, report=None) -> Iterator[dict]:
    rmax = wsv.max_row or 1
    ri = wsv.min_row or 1
    tnum = 0
    guard = 0
    prev = {}         # carries previous MULTI header for title-band reuse
    prev_single = {}  # carries previous SINGLE header for trailing-column reuse
    while ri <= rmax and guard < MAX_TABLES_PER_SHEET:
        guard += 1
        while ri <= rmax and _is_all_blank(wsv, ri, 1, cmax):
            ri += 1
        if ri > rmax:
            break
        model = _classify_header_model(wss, ri, cmax)
        if model == "multi":
            gen = _emit_multi(wsv, wss, sheet, cmax, source_file, ri, tnum, prev, report=report)
        else:
            gen = _emit_single(wsv, wss, sheet, cmax, source_file, ri, rmax, tnum, report=report,
                                prev_single=prev_single)
        next_ri = ri + 1
        for item in gen:
            if "_next_ri" in item:
                next_ri = item["_next_ri"]
            else:
                yield item
        tnum += 1
        ri = max(next_ri, ri + 1)


def iter_xlsx(path, report=None) -> Iterator[dict]:
    """Public entry point used by ingester.py. `path` is a pathlib.Path.

    If `report` is a list, it is populated in place with one meta dict per
    successfully-captured table heading (see `_emit_single`/`_emit_multi`),
    across every sheet in the workbook. Default (None) is a no-op, so existing
    callers (ingester.py) are unaffected."""
    wbv = openpyxl.load_workbook(path, data_only=True)
    wbs = load_workbook(path)
    for sheet in wbv.sheetnames:
        wsv = wbv[sheet]
        wss = wbs[sheet]
        cmax = wsv.max_column or 1
        if (wsv.max_row or 0) < 1 or cmax < 1:
            continue
        yield from _iter_sheet_tables(wsv, wss, sheet, cmax, path.name, report=report)
